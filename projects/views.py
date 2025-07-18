from django.shortcuts import render
from django.http import HttpResponse
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from django.db import models
from .models import Project, Customer
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import DimensionHolder, ColumnDimension
from io import BytesIO
from openai import OpenAI
import json

client = OpenAI()

def project_list(request):
    projects = Project.objects.select_related('customer').all()  # customerを一緒に取得する

    # 状態やキーワード絞り込みもあればここに
    status_filter = request.GET.get('status')
    keyword = request.GET.get('keyword')
    if status_filter in ['open', 'closed']:
        projects = projects.filter(status=status_filter)
    if keyword:
        projects = projects.filter(
            models.Q(customer__name__icontains=keyword) |  # customer.name で検索
            models.Q(detail__icontains=keyword)
        )

    context = {'projects': projects}
    return render(request, 'projects/project_list.html', context)


# ▼ 顧客別の案件一覧
def projects_by_customer(request, customer_id):
    customer = get_object_or_404(Customer, pk=customer_id)
    projects = Project.objects.filter(customer=customer)
    context = {
        'projects': projects,
        'customer': customer
    }
    return render(request, 'projects/project_list.html', context)

# ▼ GPTで案件情報抽出
def extract_projects_with_gpt(text):
    system_prompt = """
あなたはSES営業担当です。以下の案件一覧の文章から、複数の案件情報を抽出してください。

各案件は以下の形式のJSONにしてください：
{
  "案件名": "",
  "作業内容": "",
  "募集要件": "",
  "人数": "",
  "時期": "",
  "場所": "",
  "その他": ""
}
複数の案件がある場合はリスト形式で返してください。
"""

    user_prompt = f"以下の案件内容を解析してください：\n\n{text}"

    try:
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt}
            ],
            temperature=0.2
        )
        content = response.choices[0].message.content.strip()
        json_start = content.find("[")
        json_end = content.rfind("]") + 1
        json_str = content[json_start:json_end]
        data = json.loads(json_str)
        return data if isinstance(data, list) else [data]

    except Exception as e:
        print("GPT解析失敗:", str(e))
        return []


# ▼ Excel出力（フォーマット整形済）
def export_all_to_excel(data_list, fileobj):
    def clean(val):
        return "" if val in ["未記入", "不明", None] else str(val)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SES案件一覧"

    headers = ['No', '案件名', '作業内容', '募集要件', '人数', '時期', '場所', 'その他']
    ws.merge_cells('A1:H1')
    ws['A1'] = "WT案件一覧"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:H2')
    ws['A2'] = "※他社への無断転送、案件サイト掲載等はお控え頂ければと思います。"
    ws['A2'].alignment = Alignment(horizontal='right')
    ws['A2'].font = Font(size=8)

    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=i, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for idx, item in enumerate(data_list, 1):
        row = [
            idx,
            clean(item.get('案件名')),
            clean(item.get('作業内容')),
            clean(item.get('募集要件')),
            clean(item.get('人数')),
            clean(item.get('時期')),
            clean(item.get('場所')),
            clean(item.get('その他'))
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=3 + idx, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    dim_holder = DimensionHolder(worksheet=ws)
    for col in range(1, len(headers) + 1):
        max_len = max(len(str(ws.cell(row=row, column=col).value or '')) for row in range(1, ws.max_row + 1))
        width = min(max_len + 5, 40)
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=width)
    ws.column_dimensions.update(dim_holder)

    thin = Side(style='thin', color='000000')
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    wb.save(fileobj)

# ▼ POSTで選択された案件からGPT解析→Excel出力
@csrf_exempt
def export_selected_projects_with_gpt(request):
    if request.method == 'POST':
        selected_ids = request.POST.getlist('selected_ids')
        projects = Project.objects.filter(id__in=selected_ids)

        all_text = "\n\n".join([p.detail for p in projects if p.detail])
        print("【DEBUG】all_text:\n", all_text)   # ←ここに追加

        parsed_data = extract_projects_with_gpt(all_text)
        print("【DEBUG】parsed_data:\n", parsed_data)  # ←ここに追加

        output = BytesIO()
        export_all_to_excel(parsed_data, output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=projects.xlsx'
        return response

    return HttpResponse("POSTリクエストで送信してください", status=405)
